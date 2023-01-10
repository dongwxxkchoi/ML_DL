# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import torch
from torch import nn
from torch.utils.data import DataLoader
from torchvision import datasets
from torchvision.transforms import ToTensor, Lambda
import neural_network
import openpyxl

global i
global j

def train_loop(dataloader, model, loss_fn, optimizer):
    global i
    global j
    size = len(dataloader.dataset)
    for batch, (X, y) in enumerate(dataloader):
        # 예측(prediction)과 손실(loss) 계산
        X, y = X.to("cuda"), y.to("cuda")
        pred = model(X)
        loss = loss_fn(pred, y)

        # 역전파
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

        if batch % 100 == 0:
            loss, current = loss.item(), batch * len(X)
            ws.cell(i, j, f"loss: {loss:>7f}  [{current:>5d}/{size:>5d}]")
            wb.save("result16.xlsx")
            i += 1
            # print(f"loss: {loss:>7f}  [{current:>5d}/{size:>5d}]")


def test_loop(dataloader, model, loss_fn):
    global i
    global j
    size = len(dataloader.dataset)
    num_batches = len(dataloader)
    test_loss, correct = 0, 0

    with torch.no_grad():
        for X, y in dataloader:
            X, y = X.to("cuda"), y.to("cuda")
            pred = model(X)
            test_loss += loss_fn(pred, y).item()
            correct += (pred.argmax(1) == y).type(torch.float).sum().item()

    test_loss /= num_batches
    correct /= size
    ws.cell(i, j, f"Test Error: \n Accuracy: {(100*correct):>0.1f}%, Avg loss: {test_loss:>8f} \n")
    wb.save("result16.xlsx")
    i += 1

def initialize_weights(m):
    if hasattr(m, 'weight') and m.weight.dim() > 1:
        nn.init.kaiming_uniform(m.weight.data)

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    global i
    global j
    i = 1
    j = -3
    wb = openpyxl.Workbook()
    wb.save("result16.xlsx")
    ws = wb.active

    # learning_rates = [1e-4, 1e-3, 1e-2]
    # optimizers = list()
    #
    # batch_sizes = [256, 128, 64, 32, 16]
    # epochs = [10, 15, 20]

    learning_rates = [1e-4, 1e-3, 1e-2]
    optimizers = list()

    batch_sizes = [16]
    epochs = [30]

    training_data = datasets.FashionMNIST(
        root="data",
        train=True,
        download=True,
        transform=ToTensor()
    )

    test_data = datasets.FashionMNIST(
        root="data",
        train=False,
        download=True,
        transform=ToTensor()
    )

    dataloaders = list()

    for batch in batch_sizes:
        dataloaders.append([DataLoader(training_data, batch_size=batch), DataLoader(test_data, batch_size=batch), batch])

    model_list = list()
    model11 = neural_network.NeuralNetwork11().to("cuda")
    model_list.append(model11)
    model12 = neural_network.NeuralNetwork12().to("cuda")
    model_list.append(model12)
    model13 = neural_network.NeuralNetwork13().to("cuda")
    model_list.append(model13)
    model21 = neural_network.NeuralNetwork21().to("cuda")
    model_list.append(model21)
    model22 = neural_network.NeuralNetwork22().to("cuda")
    model_list.append(model22)
    model23 = neural_network.NeuralNetwork23().to("cuda")
    model_list.append(model23)
    model31 = neural_network.NeuralNetwork31().to("cuda")
    model_list.append(model31)
    model32 = neural_network.NeuralNetwork32().to("cuda")
    model_list.append(model32)


    for idx, model in enumerate(model_list):
        for lr in learning_rates:
            optimizers.append([torch.optim.SGD(model_list[idx].parameters(), lr=lr), "SGD", model_list[idx], lr])
            optimizers.append([torch.optim.Adam(model_list[idx].parameters(), lr=lr), "Adam", model_list[idx], lr])

    total_len = len(dataloaders) * len(learning_rates) * len(model_list)

    cur = 0

    for epoch in epochs:
        for dl_idx, dataloader in enumerate(dataloaders):
            for idx, model in enumerate(model_list):
                for lr in learning_rates:
                    print(str(cur)+","+str(total_len))
                    cur += 1
                    i = 1
                    j += 5

                    model_list[idx].apply(initialize_weights)
                    loss_fn = nn.CrossEntropyLoss()
                    optimizer = torch.optim.SGD(model_list[idx].parameters(), lr=lr)

                    ws.cell(i, j, "optimizer: SGD")
                    wb.save("result16.xlsx")
                    i += 1
                    ws.cell(i, j, "model:" + model.info)
                    wb.save("result16.xlsx")
                    i += 1
                    ws.cell(i, j, "batch:" + str(dataloader[2]))
                    wb.save("result16.xlsx")
                    i += 1
                    ws.cell(i, j, "learning rate:" + str(lr))
                    wb.save("result16.xlsx")
                    i += 1
                    ws.cell(i, j, "Epoch:" + str(epoch))
                    wb.save("result16.xlsx")
                    i += 1
                    for t in range(epoch):
                        ws.cell(i, j, f"Epoch {t + 1}\n-------------------------------")
                        wb.save("result16.xlsx")
                        i += 1
                        train_loop(dataloaders[dl_idx][0], model, loss_fn, optimizer)
                        test_loop(dataloaders[dl_idx][1], model, loss_fn)
                    print("Done!")

    # for epoch in epochs:
    #     for dl_idx, dataloader in enumerate(dataloaders):
    #         for idx, model in enumerate(model_list):
    #             for lr in learning_rates:
    #                 model_list[idx].apply(initialize_weights)
    #                 loss_fn = nn.CrossEntropyLoss()
    #                 optimizer = torch.optim.Adam(model_list[idx].parameters(), lr=lr)
    #                 print("optimizer: Adam")
    #                 print("model:", model.info)
    #                 print("batch:", dataloader[2])
    #                 print("learning rate:", lr)
    #                 print("Epoch:", epoch)
    #                 for t in range(epoch):
    #                     print(f"Epoch {t + 1}\n-------------------------------")
    #                     train_loop(dataloaders[dl_idx][0], model, loss_fn, optimizer)
    #                     test_loop(dataloaders[dl_idx][1], model, loss_fn)
    #                 print("Done!")

    # for epoch in epochs:
    #     for dl_idx, dataloader in enumerate(dataloaders):
    #         for opt_idx, optimizer in enumerate(optimizers):
    #             loss_fn = nn.CrossEntropyLoss()
    #             print("optimizer:", optimizer[1])
    #             print("model:", optimizer[2].info)
    #             print("batch:", dataloader[2])
    #             print("learning rate:", optimizer[3])
    #             print("Epoch:", epoch)
    #             for t in range(epoch):
    #                 print(f"Epoch {t + 1}\n-------------------------------")
    #                 train_loop(dataloaders[dl_idx][0], optimizers[opt_idx][2], loss_fn, optimizers[opt_idx][0])
    #                 test_loop(dataloaders[dl_idx][1], optimizers[opt_idx][2], loss_fn)
    #             print("Done!")
    #             loss_fn = None

    # for opt_idx, optimizer in enumerate(optimizers):
    #     for epoch in epochs:
    #         for dl_idx, dataloader in enumerate(dataloaders):
    #             print("optimizer:", optimizer[1])
    #             print("model:", optimizer[2].info)
    #             print("batch:", dataloader[2])
    #             print("learning rate:", optimizer[3])
    #             print("Epoch:", epoch)
    #             for t in range(epoch):
    #                 print(f"Epoch {t + 1}\n-------------------------------")
    #                 train_loop(dataloaders[dl_idx][0], optimizers[opt_idx][2], loss_fn, optimizers[opt_idx][0])
    #                 test_loop(dataloaders[dl_idx][1], optimizers[opt_idx][2], loss_fn)
    #             print("Done!")